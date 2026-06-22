from django.views.generic import ListView, TemplateView
from .models import ProfesorCursoAsignatura, Curso, Asignatura, Estudiante, Aviso, Actividad, Calificacion, EstudianteCurso, Usuario
from django.http import JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
import json
from .login_mixin import LoginRequeridoMixin
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse


# Create your views here.
class ListaCursosView(LoginRequeridoMixin, ListView):
    model = Curso
    template_name = 'academico/lista_cursos.html'  
    context_object_name = 'cursos_asignados'

    def get_queryset(self):
        profesor_id = self.request.session.get('usuario_id')
        
        # Agregamos select_related para precargar la información de la asignatura adjunta
        asignaciones = ProfesorCursoAsignatura.objects.filter(id_profesor=profesor_id).select_related('id_asignatura')
        # distinct elimina duplicados
        curso_ids = asignaciones.values_list('id_curso', flat=True).distinct()
        
        cursos = Curso.objects.filter(id_curso__in=curso_ids)
        for curso in cursos:
            curso.materias_del_profesor = asignaciones.filter(id_curso=curso.id_curso)
            
        return cursos

class ListaMateriasCursoView(LoginRequeridoMixin, ListView):
    model = ProfesorCursoAsignatura
    template_name = 'academico/lista_materias.html'
    context_object_name = 'materias_asignadas'
    
    def get_queryset(self):
        profesor_id = self.request.session.get('usuario_id')
        id_del_curso = self.kwargs['curso_id']
        
        asignaciones = ProfesorCursoAsignatura.objects.filter(
            id_profesor = profesor_id,
            id_curso = id_del_curso
        )
        
        asignatura_ids = asignaciones.values_list('id_asignatura', flat=True)
        return Asignatura.objects.filter(id_asignatura__in=asignatura_ids)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cursos'] = Curso.objects.get(pk=self.kwargs['curso_id'])
        return context
        
        
class PanelMateriaView(LoginRequeridoMixin, TemplateView):
    template_name = 'academico/panel_materia.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        id_del_curso = self.kwargs['curso_id']
        id_de_la_asignatura = self.kwargs['asignatura_id']
        
        context['curso'] = Curso.objects.get(pk=id_del_curso)
        context['asignatura'] = Asignatura.objects.get(pk=id_de_la_asignatura)
        
        colegio = context['curso'].colegio
        context['nota_minima'] = colegio.nota_minima
        context['nota_maxima'] = colegio.nota_maxima
        
        pca = ProfesorCursoAsignatura.objects.filter(
            id_curso=id_del_curso,
            id_asignatura=id_de_la_asignatura
        ).first()
        
        # Pestaña 1: Avisos
        context['publicaciones'] = Aviso.objects.filter(
            curso_id=id_del_curso,       
            asignatura_id=id_de_la_asignatura  
        ).select_related('autor').order_by('-fecha_publicacion')

        # Pestaña 2: Actividades
        if pca:
            context['asignaciones'] = Actividad.objects.filter(id_pca=pca).order_by('-fecha_limite')
        else:
            context['asignaciones'] = Actividad.objects.none()

        # Pestaña 4: Alumnos
        matriculas = EstudianteCurso.objects.filter(
            id_curso=id_del_curso,
            estado='inscrito'
        ).select_related('id_estudiante__id_estudiante')
    
        
        estudiantes = [m.id_estudiante for m in matriculas]
        context['estudiantes'] = estudiantes

        # Pestaña 3: Planilla de Calificaciones (NUEVA LÓGICA)
        if pca:
            calificaciones_qs = Calificacion.objects.filter(id_actividad__id_pca=pca)
        else:
            calificaciones_qs = Calificacion.objects.none()
        
        # Juntamos las notas en un diccionario rápido usando tuplas de (alumno_id, actividad_id)
        notas_dict = {(c.id_estudiante_id, c.id_actividad_id): c.nota for c in calificaciones_qs}

        # Cruzamos los alumnos con sus notas de una vez
        planilla_alumnos = []
        for est in estudiantes:
            filas_actividades = []
            for trab in context['asignaciones']:
                nota_actual = notas_dict.get((est.id_estudiante_id, trab.id_actividad), '')
                filas_actividades.append({
                    'actividad_id': trab.id_actividad,
                    'nota': nota_actual
                })
            planilla_alumnos.append({
                'estudiante': est,
                'actividades': filas_actividades
            })
        
        context['planilla_alumnos'] = planilla_alumnos

        return context
    
    
@method_decorator(csrf_exempt, name='dispatch')
class CrearActividadView(View):
    def post(self, request, curso_id, asignatura_id):
        try:
            data = json.loads(request.body)
            
            pca = ProfesorCursoAsignatura.objects.filter(
                id_curso=curso_id,
                id_asignatura=asignatura_id
            ).first()
            
            if not pca:
                return JsonResponse({'error': 'No se encontró la asignación'}, status=404)
            
            actividad = Actividad.objects.create(
                nombre=data.get('nombre'),
                descripcion=data.get('descripcion', ''),
                fecha_limite=data.get('fecha_limite') or None,
                id_pca=pca
            )
            
            return JsonResponse({
                'ok': True,
                'id': actividad.id_actividad,
                'nombre': actividad.nombre,
                'fecha_limite': str(actividad.fecha_limite) if actividad.fecha_limite else None
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        
        

@method_decorator(csrf_exempt, name='dispatch')
class EditarActividadView(View):
    def post(self, request, curso_id, asignatura_id, actividad_id):
        try:
            data = json.loads(request.body)
            actividad = Actividad.objects.get(pk=actividad_id)
            actividad.nombre = data.get('nombre', actividad.nombre)
            actividad.descripcion = data.get('descripcion', actividad.descripcion)
            actividad.fecha_limite = data.get('fecha_limite') or None
            actividad.save()
            return JsonResponse({
                'ok': True,
                'id': actividad.id_actividad,
                'nombre': actividad.nombre,
                'fecha_limite': str(actividad.fecha_limite) if actividad.fecha_limite else None
            })
        except Actividad.DoesNotExist:
            return JsonResponse({'error': 'Actividad no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        
        
@method_decorator(csrf_exempt, name='dispatch')
class EliminarActividadView(View):
    def post(self, request, curso_id, asignatura_id, actividad_id):
        try:
            actividad = Actividad.objects.get(pk=actividad_id)
            actividad.delete()
            return JsonResponse({'ok': True})
        except Actividad.DoesNotExist:
            return JsonResponse({'error': 'Actividad no encontrada'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        
        
        
@method_decorator(csrf_exempt, name='dispatch')
class CrearAvisoView(View):
    def post(self, request, curso_id, asignatura_id):
        try:
            data = json.loads(request.body)
            contenido = data.get('contenido', '').strip()

            if not contenido:
                return JsonResponse({'error': 'El contenido no puede estar vacío'}, status=400)

            profesor_id = 21  # mismo id fijo usado en el resto de vistas
            autor = Usuario.objects.get(pk=profesor_id)

            aviso = Aviso.objects.create(
                autor=autor,
                contenido=contenido,
                curso_id=curso_id,
                asignatura_id=asignatura_id
            )

            return JsonResponse({
                'ok': True,
                'id': aviso.id_aviso,
                'contenido': aviso.contenido,
                'autor_nombre': autor.primer_nombre,
                'autor_apellido': autor.primer_apellido,
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class GuardarNotaView(View):
    def post(self, request, curso_id, asignatura_id):
        try:
            data = json.loads(request.body)
            estudiante_id = data.get('estudiante_id')
            actividad_id = data.get('actividad_id')
            nota = data.get('nota')

            nota_valor = nota if nota not in (None, '') else None

            if nota_valor is not None:
                curso = Curso.objects.select_related('colegio').get(pk=curso_id)
                nota_min = curso.colegio.nota_minima
                nota_max = curso.colegio.nota_maxima

                nota_valor = float(nota_valor)
                if nota_valor < float(nota_min) or nota_valor > float(nota_max):
                    return JsonResponse({
                        'error': f'La nota debe estar entre {nota_min} y {nota_max}'
                    }, status=400)

            calificacion, created = Calificacion.objects.update_or_create(
                id_estudiante_id=estudiante_id,
                id_actividad_id=actividad_id,
                defaults={
                    'nota': nota_valor,
                    'colegio_id': 1,
                    'estado': 'entregado' if nota_valor is not None else 'sin_entregar',
                }
            )

            return JsonResponse({'ok': True, 'nota': str(calificacion.nota) if calificacion.nota is not None else None})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
        
        
class ExportarNotasView(LoginRequeridoMixin, View):
    def get(self, request, curso_id, asignatura_id):
        try:
            curso = Curso.objects.get(pk=curso_id)
            asignatura = Asignatura.objects.get(pk=asignatura_id)

            pca = ProfesorCursoAsignatura.objects.filter(
                id_curso=curso_id,
                id_asignatura=asignatura_id
            ).first()

            asignaciones = Actividad.objects.filter(id_pca=pca).order_by('fecha_limite') if pca else Actividad.objects.none()

            matriculas = EstudianteCurso.objects.filter(
                id_curso=curso_id,
                estado='inscrito'
            ).select_related('id_estudiante__id_estudiante')
            estudiantes = [m.id_estudiante for m in matriculas]

            calificaciones_qs = Calificacion.objects.filter(id_actividad__id_pca=pca) if pca else Calificacion.objects.none()
            notas_dict = {(c.id_estudiante_id, c.id_actividad_id): c.nota for c in calificaciones_qs}

            # ── Construcción del Excel ──
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Calificaciones"

            # Encabezado
            ws['A1'] = f"Planilla de Calificaciones - {curso.nombre_curso} - {asignatura.nombre_asignatura}"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + asignaciones.count())
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Fila de columnas (fila 3)
            fila_header = 3
            ws.cell(row=fila_header, column=1, value="Estudiante").font = Font(bold=True)
            ws.cell(row=fila_header, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            for idx, act in enumerate(asignaciones, start=2):
                celda = ws.cell(row=fila_header, column=idx, value=act.nombre)
                celda.font = Font(bold=True)
                celda.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                celda.alignment = Alignment(horizontal='center')

            # Filas de estudiantes
            fila_actual = fila_header + 1
            for est in estudiantes:
                usuario = est.id_estudiante
                nombre_completo = f"{usuario.primer_nombre} {usuario.primer_apellido}"
                ws.cell(row=fila_actual, column=1, value=nombre_completo)

                for idx, act in enumerate(asignaciones, start=2):
                    nota = notas_dict.get((est.id_estudiante_id, act.id_actividad), '')
                    celda = ws.cell(row=fila_actual, column=idx, value=float(nota) if nota != '' and nota is not None else '')
                    celda.alignment = Alignment(horizontal='center')

                fila_actual += 1

            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 30
            for idx in range(2, 2 + asignaciones.count()):
                ws.column_dimensions[ws.cell(row=fila_header, column=idx).column_letter].width = 18

            # ── Respuesta HTTP como archivo descargable ──
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            nombre_archivo = f"calificaciones_{curso.nombre_curso}_{asignatura.nombre_asignatura}.xlsx".replace(' ', '_')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            wb.save(response)

            return response

        except Exception as e:
            print(f"DEBUG: Error en ExportarNotasView - {e}")
            return HttpResponse(f"Error al generar el Excel: {e}", status=500)